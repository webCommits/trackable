"""Parser for time-entry import files (CSV, XLSX, XLS, ODS)."""

import csv
import io
import os
import re
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from typing import Optional


class ParsedEntry:
    """A parsed time entry from a spreadsheet row."""

    def __init__(
        self,
        date: date,
        start_time: Optional[time] = None,
        end_time: Optional[time] = None,
        notes: str = "",
        duration: Optional[Decimal] = None,
        source_sheet: str = "",
        source_row: int = 0,
    ):
        self.date = date
        self.start_time = start_time
        self.end_time = end_time
        self.notes = notes
        self.duration = duration
        self.source_sheet = source_sheet
        self.source_row = source_row

    def __repr__(self):
        return (
            f"ParsedEntry(date={self.date}, start={self.start_time}, "
            f"end={self.end_time}, notes={self.notes!r}, duration={self.duration})"
        )


def detect_file_format(filename: str) -> str:
    """Detect file format from extension.

    Returns one of: 'csv', 'xlsx', 'xls', 'ods'.

    Raises ValueError if the extension is unknown.
    """
    ext = os.path.splitext(filename)[1].lower()
    mapping = {
        ".csv": "csv",
        ".xlsx": "xlsx",
        ".xls": "xls",
        ".ods": "ods",
    }
    if ext in mapping:
        return mapping[ext]
    raise ValueError(
        f"Unsupported file format '{ext}'. Supported: .csv, .xlsx, .xls, .ods"
    )


def _parse_date(value: str, fmt: str) -> Optional[date]:
    """Try to parse a date string with the given format."""
    if not value or not value.strip():
        return None
    try:
        return datetime.strptime(value.strip(), fmt).date()
    except ValueError:
        return None


def _parse_time(value: str, fmt: str) -> Optional[time]:
    """Try to parse a time string with the given format."""
    if not value or not value.strip():
        return None
    try:
        return datetime.strptime(value.strip(), fmt).time()
    except ValueError:
        return None


def _parse_decimal(value: str, decimal_separator: str) -> Optional[Decimal]:
    """Parse a decimal number, handling the given decimal separator."""
    if not value or not value.strip():
        return None
    cleaned = value.strip().replace(" ", "")
    if decimal_separator == ",":
        cleaned = cleaned.replace(",", ".")
    # Remove thousand separators (dots or commas depending on locale)
    # After replacing decimal sep, handle thousand sep
    if decimal_separator == ",":
        # thousand sep is dot: 1.000,50 -> 1.000.50 -> replace dots -> 1000.50
        # But only dots that are thousand separators
        parts = cleaned.split(".")
        if len(parts) > 2:
            # Could be thousand separator: 1.000.50
            # If last part has <= 2 digits, it's decimal
            # We already replaced comma, so just remove all dots
            cleaned = cleaned.replace(".", "")
        # else: single dot is already the decimal we converted
    else:
        # decimal sep is ., thousand sep is comma
        cleaned = cleaned.replace(",", "")
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


def read_spreadsheet(
    file,
    file_format: str,
    import_all_sheets: bool = True,
    sheet_name: Optional[str] = None,
    separator: Optional[str] = None,
) -> list[tuple[str, list[list[str]]]]:
    """Read a spreadsheet file and return (sheet_name, rows) pairs.

    Each row is a list of string cell values.
    """
    if file_format == "csv":
        return _read_csv(file, separator=separator)
    elif file_format == "xlsx":
        return _read_xlsx(file, import_all_sheets, sheet_name)
    elif file_format == "xls":
        return _read_xls(file, import_all_sheets, sheet_name)
    elif file_format == "ods":
        return _read_ods(file, import_all_sheets, sheet_name)
    else:
        raise ValueError(f"Unsupported file format: {file_format}")


def _read_csv(file, separator: Optional[str] = None) -> list[tuple[str, list[list[str]]]]:
    """Read a CSV file. Uses chardet for encoding detection, then csv module.

    If separator is provided, it is used as the CSV delimiter.
    Otherwise, csv.Sniffer() is used to auto-detect.
    """
    raw = file.read()
    if isinstance(raw, str):
        raw = raw.encode("utf-8")

    import chardet

    detected = chardet.detect(raw)
    encoding = detected.get("encoding", "utf-8") or "utf-8"
    # Handle BOM
    if encoding and encoding.upper() in ("UTF-8-SIG",):
        encoding = "utf-8-sig"

    text = raw.decode(encoding, errors="replace")

    if separator:
        delimiter = separator
    else:
        # Try to detect delimiter
        try:
            sample = text[:4096]
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";"

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader]

    return [("CSV", rows)]


def _read_xlsx(
    file, import_all_sheets: bool, sheet_name: Optional[str]
) -> list[tuple[str, list[list[str]]]]:
    """Read an .xlsx file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    return _extract_sheets(wb, import_all_sheets, sheet_name)


def _read_xls(
    file, import_all_sheets: bool, sheet_name: Optional[str]
) -> list[tuple[str, list[list[str]]]]:
    """Read an .xls file using xlrd."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=file.read())
    result = []
    for i in range(wb.nsheets):
        ws = wb.sheet_by_index(i)
        name = ws.name
        if not import_all_sheets and sheet_name and name != sheet_name:
            continue
        rows = []
        for row_idx in range(ws.nrows):
            row_vals = [
                str(ws.cell_value(row_idx, col_idx)) if ws.cell_type(row_idx, col_idx) != xlrd.XL_CELL_EMPTY else ""
                for col_idx in range(ws.ncols)
            ]
            rows.append(row_vals)
        result.append((name, rows))
    return result


def _read_ods(
    file, import_all_sheets: bool, sheet_name: Optional[str]
) -> list[tuple[str, list[list[str]]]]:
    """Read an .ods file using odfpy."""
    from odf import opendocument, table
    from lxml import etree

    doc = opendocument.load(file)
    ns = {"table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
          "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0"}
    # doc.xml is a method that returns bytes, parse with lxml
    xml_bytes = doc.xml()
    root = etree.fromstring(xml_bytes)
    tables = root.findall(".//table:table", ns)

    result = []
    for tbl in tables:
        name = tbl.get("{urn:oasis:names:tc:opendocument:xmlns:table:1.0}name")
        if not import_all_sheets and sheet_name and name != sheet_name:
            continue
        rows = tbl.findall("table:table-row", ns)
        sheet_rows = []
        for row in rows:
            cells = row.findall("table:table-cell", ns)
            row_vals = []
            for cell in cells:
                # Check for repeated columns
                repeat = cell.get(
                    "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}number-columns-repeated"
                )
                repeat = int(repeat) if repeat else 1
                text_parts = cell.findall("text:p", ns) if repeat == 1 else []
                txt = "".join(p.text or "" for p in text_parts).strip()
                row_vals.append(txt)
                for _ in range(repeat - 1):
                    row_vals.append("")
            sheet_rows.append(row_vals)

        if import_all_sheets or (sheet_name and name == sheet_name):
            result.append((name, sheet_rows))

    return result


def _extract_sheets(wb, import_all_sheets, sheet_name):
    """Extract all or specific sheets from an openpyxl workbook."""
    result = []
    for ws in wb.worksheets:
        if not import_all_sheets and sheet_name and ws.title != sheet_name:
            continue
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 0):
            row_vals = [
                str(cell.value) if cell.value is not None else "" for cell in row
            ]
            rows.append(row_vals)
        result.append((ws.title, rows))
    return result


def find_header_row(rows: list[list[str]]) -> int:
    """Find the row index containing 'Datum' in the first column.

    Returns the 0-based index. Raises ValueError if not found.
    """
    for i, row in enumerate(rows):
        if row and len(row) > 0:
            first_cell = row[0].strip().lower()
            if first_cell == "datum":
                return i
    raise ValueError(
        "Could not find header row containing 'Datum' in the first column."
    )


def is_summary_row(row: list[str]) -> bool:
    """Return True if the row contains summary keywords."""
    summary_keywords = {"ist", "soll", "übertrag", "über-/unterdeckung"}
    for cell in row:
        # Normalize: strip and remove internal spaces for matching
        cell_lower = cell.strip().lower().replace(" ", "")
        if cell_lower in summary_keywords:
            return True
        # Also check if cell starts with a keyword (e.g. "Übertrag aus Vormonat")
        for kw in summary_keywords:
            if cell_lower.startswith(kw):
                return True
    # Also check first non-empty cell
    for cell in row:
        if cell.strip():
            cell_lower = cell.strip().lower().replace(" ", "")
            if cell_lower in summary_keywords:
                return True
            for kw in summary_keywords:
                if cell_lower.startswith(kw):
                    return True
            break
    return False


def _extract_sheet_year(sheet_name: str) -> Optional[int]:
    """Try to extract a 4-digit year from a sheet name like 'Januar 2026'."""
    match = re.search(r"\b(20\d{2})\b", sheet_name)
    if match:
        return int(match.group(1))
    return None


def parse_rows(
    rows: list[list[str]],
    header_row: int,
    config: Optional[dict] = None,
    sheet_name: str = "",
) -> tuple[list[ParsedEntry], list[str], list[str]]:
    """Parse data rows after a header row into ParsedEntry objects.

    Returns (entries, warnings, errors).
    """
    if config is None:
        config = {}

    date_col = config.get("date_col", 0)
    start_col = config.get("start_col", 1)
    end_col = config.get("end_col", 2)
    duration_col = config.get("duration_col", 3)
    notes_col = config.get("notes_col", 4)
    decimal_separator = config.get("decimal_separator", ",")
    date_format = config.get("date_format", "%d.%m.%Y")
    time_format = config.get("time_format", "%H:%M")

    entries: list[ParsedEntry] = []
    warnings: list[str] = []
    errors: list[str] = []

    sheet_year = _extract_sheet_year(sheet_name)

    for i, row in enumerate(rows):
        row_idx = header_row + 1 + i
        if row_idx >= len(rows):
            break
        row_data = rows[row_idx]

        # Check for empty row
        if all(_is_cell_empty(cell) for cell in row_data):
            continue

        # Check for summary row
        if is_summary_row(row_data):
            continue

        date_val = _get_cell(row_data, date_col)
        start_val = _get_cell(row_data, start_col)
        end_val = _get_cell(row_data, end_col)
        duration_val = _get_cell(row_data, duration_col)
        notes_val = _get_cell(row_data, notes_col)

        # Skip if all relevant columns are empty
        relevant = [date_val, start_val, end_val, duration_val, notes_val]
        if all(not v for v in relevant):
            continue

        # Parse date
        parsed_date = _parse_date(date_val, date_format)
        if parsed_date is None:
            errors.append(
                _row_error(
                    row_idx,
                    f"Invalid date '{date_val}' (expected format: {date_format})",
                    sheet_name,
                )
            )
            continue

        # Check date year vs sheet name year
        if sheet_year is not None and parsed_date.year != sheet_year:
            warnings.append(
                _row_warning(
                    row_idx,
                    f"Date year ({parsed_date.year}) does not match sheet year "
                    f"({sheet_year})",
                    sheet_name,
                )
            )

        # Determine if we have start/end times or just duration
        has_start_end = bool(start_val.strip()) and bool(end_val.strip())

        if has_start_end:
            start_time = _parse_time(start_val, time_format)
            end_time = _parse_time(end_val, time_format)
            if start_time is None or end_time is None:
                errors.append(
                    _row_error(
                        row_idx,
                        f"Invalid time format: start='{start_val}', end='{end_val}'",
                        sheet_name,
                    )
                )
                continue

            # Duration comparison
            duration_decimal = _parse_decimal(duration_val, decimal_separator)
            if duration_decimal is not None:
                # Calculate duration from start/end
                from datetime import datetime, timedelta

                start_dt = datetime.combine(parsed_date, start_time)
                end_dt = datetime.combine(parsed_date, end_time)
                if end_dt < start_dt:
                    end_dt += timedelta(days=1)
                actual_hours = Decimal(str(
                    round((end_dt - start_dt).total_seconds() / 3600, 2)
                ))
                diff = abs(actual_hours - duration_decimal)
                if diff > Decimal("0.25"):
                    warnings.append(
                        _row_warning(
                            row_idx,
                            f"Computed duration ({actual_hours}h) differs from "
                            f"file's duration ({duration_decimal}h) by {diff}h",
                            sheet_name,
                        )
                    )

            entries.append(
                ParsedEntry(
                    date=parsed_date,
                    start_time=start_time,
                    end_time=end_time,
                    notes=notes_val,
                    source_sheet=sheet_name,
                    source_row=row_idx,
                )
            )

        elif duration_val.strip():
            # Stundengutschrift: no start/end, but duration
            duration_decimal = _parse_decimal(duration_val, decimal_separator)
            if duration_decimal is None:
                errors.append(
                    _row_error(
                        row_idx,
                        f"Invalid duration value '{duration_val}'",
                        sheet_name,
                    )
                )
                continue

            if duration_decimal >= Decimal("24"):
                errors.append(
                    _row_error(
                        row_idx,
                        f"Duration {duration_decimal}h is >= 24h, cannot import",
                        sheet_name,
                    )
                )
                continue

            # Create entry with start_time=00:00 and end_time = duration hours later
            start_time = time(0, 0)
            end_minutes = int(duration_decimal * 60)
            end_hour = end_minutes // 60
            end_min = end_minutes % 60
            if end_hour >= 24:
                end_hour = 23
                end_min = 59
            end_time = time(end_hour, end_min)

            # Add note about Stundengutschrift
            combined_notes = f"{notes_val.strip()} [Stundengutschrift: {duration_decimal}h]" if notes_val.strip() else f"Stundengutschrift: {duration_decimal}h"

            entries.append(
                ParsedEntry(
                    date=parsed_date,
                    start_time=start_time,
                    end_time=end_time,
                    notes=combined_notes,
                    duration=duration_decimal,
                    source_sheet=sheet_name,
                    source_row=row_idx,
                )
            )

        else:
            warnings.append(
                _row_warning(
                    row_idx,
                    "Row has no start/end times and no duration - skipped",
                    sheet_name,
                )
            )

    return entries, warnings, errors


def _is_cell_empty(value: str) -> bool:
    """Check if a cell value is effectively empty."""
    return not value.strip()


def _get_cell(row: list[str], col: int) -> str:
    """Get cell value at column index, returning empty string if out of range."""
    if col < len(row):
        return row[col].strip()
    return ""


def _row_warning(row_idx: int, message: str, sheet_name: str = "") -> str:
    """Format a row-level warning message."""
    prefix = f"[{sheet_name}:{row_idx + 1}]" if sheet_name else f"[Row {row_idx + 1}]"
    return f"{prefix} {message}"


def _row_error(row_idx: int, message: str, sheet_name: str = "") -> str:
    """Format a row-level error message."""
    prefix = f"[{sheet_name}:{row_idx + 1}]" if sheet_name else f"[Row {row_idx + 1}]"
    return f"{prefix} {message}"
